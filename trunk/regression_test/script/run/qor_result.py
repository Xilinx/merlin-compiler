#(C) Copyright 2016-2021 Xilinx, Inc.
#All Rights Reserved.
#
#Licensed to the Apache Software Foundation (ASF) under one
#or more contributor license agreements.  See the NOTICE file
#distributed with this work for additional information
#regarding copyright ownership.  The ASF licenses this file
#to you under the Apache License, Version 2.0 (the
#"License"); you may not use this file except in compliance
#with the License.  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#Unless required by applicable law or agreed to in writing,
#software distributed under the License is distributed on an
#"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
#KIND, either express or implied.  See the License for the
#specific language governing permissions and limitations
#under the License. (edited)
#!/usr/bin/python

import sys
import os
import fnmatch
import re
import numbers
import time
from datetime import datetime, date

import csv
from openpyxl import Workbook
from openpyxl import load_workbook

curr_date = datetime.today()

run_sim_profile = ""
run_bit_profile = ""
curr_fpga_design = ""
resource_report = ""
resource_report2 = ""
impl_report = ""
impl_report2 = ""
hls_report = ""
hls_report2 = ""
hls_report3 = ""
merlin_report = ""
merlin_report2 = ""

show_hls = True
show_impl = True
show_onboard = True

def run_bit_profile_setup(dir):
    # get profil16yye file path
    global run_bit_profile, curr_fpga_design
    for root,dirnames,filenames in os.walk(dir) :
        if "run_bit" in root.split("/") :
            #print (root)
            #print (filenames)
            if "sdaccel_profile_summary.csv" in filenames :
                #print (os.path.join(root, "sdaccel_profile_summary.csv"))
                run_bit_profile = os.path.join(root, "sdaccel_profile_summary.csv")
                curr_fpga_design = "x"
                break
            elif "profile.mon" in filenames :
                #print (os.path.join(root, "profile.mon"))
                run_bit_profile = os.path.join(root, "profile.mon")
                curr_fpga_design = "i"
                break

def run_sim_profile_setup(dir):
    # get profile file path
    global run_sim_profile, curr_fpga_design
    for root,dirnames,filenames in os.walk(dir) :
        if "run_sim" in root.split("/") :
            #print (root)
            #print (filenames)
            if "sdaccel_profile_summary.csv" in filenames :
                #print (os.path.join(root, "sdaccel_profile_summary.csv"))
                run_sim_profile = os.path.join(root, "sdaccel_profile_summary.csv")
                curr_fpga_design = "x"
                break
            elif "profile.mon" in filenames :
                #print (os.path.join(root, "profile.mon"))
                run_sim_profile = os.path.join(root, "profile.mon")
                curr_fpga_design = "i"
                break

def run_bit_profile_x_parsing(profile_name):
    bs_kernel = 0.0
    mem_transfer = 0.0
    #print ("x:" + profile_name)
    if os.path.isfile(profile_name) == False :
        return (bs_kernel, mem_transfer)
    f = open(profile_name, 'r')
    while True :
        line = f.readline()
        if line == "" :
            break
        else :
            data1 = line.rstrip('\n').split(",")
            if len(data1) == 1 and data1[0] != "" :
                if data1[0] == "Kernel Execution" :
                    f.readline()
                    line = f.readline()
                    data2 = line.rstrip('\n').split(",")
                    if len(data2) > 2 :
                        bs_kernel = float(data2[2])
                    else :
                        bs_kernel = 0.0
                elif data1[0] == "Data Transfer: Host and Global Memory" :
                    f.readline()
                    line = f.readline()
                    data2 = line.rstrip('\n').split(",")
                    line = f.readline()
                    data3 = line.rstrip('\n').split(",")
                    if len(data2) > 6 and len(data3) > 6 :
                        mem_transfer = (float(data2[6]) + float(data3[6]))
                    else :
                        mem_transfer = 0.0
    f.close()
    #print ("Kernel execution time=" + str(bs_kernel) + " s")
    #print ("PCIe transfer time=" + str(mem_transfer) + " s")
    return (bs_kernel, mem_transfer)

def run_bit_profile_i_parsing(profile_name):
    mem_transfer = 0.0
    bs_kernel = 0.0
    last_start = 0.0
    last_end = 0.0
    #print ("i:" + profile_name)
    if os.path.isfile(profile_name) == False :
        return (bs_kernel, mem_transfer)
    with open(profile_name) as file:
        lineList = file.readlines()
        for line in lineList:
            data = line.rstrip('\n').split(",")
            if data[2] == ".mem_transfer" :
                transfer = (float(data[4]) - float(data[3])) / 1000000.0
                overlap = min(float(last_end), float(data[4])) - max(float(last_start), float(data[3]))
                mem_transfer += (float(data[4]) - float(data[3])) / 1000000.0
                if overlap > 0.0 :
                    mem_transfer -= overlap / 1000000000.0
            elif re.search('.*_kernel', data[2]) != None or re.search('workload', data[2]) != None :
                bs_kernel += (float(data[4]) - float(data[3])) / 1000000.0
            last_start = data[3]
            last_end = data[4]
    #print ("Kernel execution time=" + str(bs_kernel) + " s")
    #print ("PCIe transfer time=" + str(mem_transfer) + " s")
    return (bs_kernel, mem_transfer)

def run_bit_profile_none():
    mem_transfer = 0.0
    bs_kernel = 0.0
    return (bs_kernel, mem_transfer)

def runaccexe_profile_setup(dir):
    # get profil16yye file path
    global run_bit_profile, curr_fpga_design
    retval = False
    #print (dir)
    run_bit_profile = ""
    for root,dirnames,filenames in os.walk(dir) :
        if "sdaccel_profile_summary.csv" in filenames :
            #print (os.path.join(root, "sdaccel_profile_summary.csv"))
            run_bit_profile = os.path.join(root, "sdaccel_profile_summary.csv")
            curr_fpga_design = "x"
            retval = True
            break
        elif "profile.mon" in filenames :
            #print (os.path.join(root, "profile.mon"))
            run_bit_profile = os.path.join(root, "profile.mon")
            curr_fpga_design = "i"
            retval = True
            break
    return (retval)

def resource_report_setup(dir):
    # get resource file path
    global resource_report, resource_report2, curr_fpga_design
    global hls_report, hls_report2
    global merlin_report
    resource_report = ""
    hls_report = ""
    hls_report2 = ""
    curr_fpga_design = ""
    for root,dirnames,filenames in os.walk(dir) :
        if "summary_impl_.rpt" in filenames :
            resource_report = os.path.join(root, "summary_impl_.rpt")
            curr_fpga_design = "x"
            #print (resource_report)
        elif "summary.log" in filenames and "acl_quartus_report.txt" in filenames :
            resource_report = os.path.join(root, "summary.log")
            resource_report2 = os.path.join(root, "acl_quartus_report.txt")
            curr_fpga_design = "i"
            #print (resource_report)
            #print (resource_report2)
        elif "merlin.rpt" in filenames and "merlin_hls.log" in filenames :
            hls_report = os.path.join(root, "merlin.rpt")
            hls_report2 = os.path.join(root, "merlin_hls.log")
            #print (hls_report)
            #print (hls_report2)
        elif "merlin.log" in filenames :
            merlin_report = os.path.join(root, "merlin.log")
            #print (merlin_report)
    #print ("~" + resource_report + "~")
    #print ("~" + curr_fpga_design + "~")

def resource_report_x_parsing(resource_name):
    lut_usage = 0
    ff_usage = 0
    ram_usage = 0
    dsp_usage = 0
    fmax_usage = 0.0
    if os.path.isfile(resource_name) == False :
        return (lut_usage, ff_usage, ram_usage, dsp_usage, fmax_usage)
    #print ("x:" +  resource_name)
    with open(resource_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('.*_kernel ', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split("|")
                lut_usage = data1[2].replace(" ", "")
                #print (lut_usage)
                ff_usage = data1[4].replace(" ", "")
                #print (ff_usage)
                ram_usage = data1[5].replace(" ", "")
                #print (ram_usage)
                dsp_usage = data1[6].replace(" ", "")
                #print (dsp_usage)
            elif re.search('Kernel Frequency', line) != None :
                #print (line)
                data1 = line.split()
                fmax_usage = float(data1[2].strip())
                #print (str(fmax_usage))
    return (lut_usage, ff_usage, ram_usage, dsp_usage, fmax_usage)

def resource_report_i_parsing(resource_name, resource_name2):
    lut_usage = 0
    ff_usage = 0
    ram_usage = 0
    dsp_usage = 0
    fmax_usage = 0.0
    if os.path.isfile(resource_name) == False :
        return (lut_usage, ff_usage, ram_usage, dsp_usage, fmax_usage)
    #print ("i:" +  resource_name)
    with open(resource_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('Total', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split("|")
                lut_usage = data1[2].replace(" ", "")
                #print (lut_usage)
                ff_usage = data1[3].replace(" ", "")
                #print (ff_usage)
                ram_usage = data1[4].replace(" ", "")
                #print (ram_usage)
                dsp_usage = data1[5].replace(" ", "")
                #print (dsp_usage)

    with open(resource_name2) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('Kernel fmax', line) != None :
                #print (line)
                data1 = line.split(":")
                fmax_usage = float(data1[1].strip())
                #print (str(fmax_usage))
    return (lut_usage, ff_usage, ram_usage, dsp_usage, fmax_usage)
    
def impl_report_x_parsing(resource_name):
    lut_usage = ""
    ff_usage = ""
    ram_usage = ""
    dsp_usage = ""
    lut_percent = ""
    ff_percent = ""
    ram_percent = ""
    dsp_percent = ""
    fmax_usage = 0.0
    if os.path.isfile(resource_name) == False :
        return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, fmax_usage)
    #print ("x:" +  resource_name)
    with open(resource_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if (re.search('.*_kernel ', line) != None or re.search('workload ', line) != None) and re.search('%', line) and re.search('\[', line) != None :
                #print (line)
                data1 = line.split("|")
                data2 = data1[2].split("[")
                lut_usage = data2[0].strip()
                lut_percent = data2[1].strip(" %]")
                #print (lut_usage)
                #print (lut_percent)
                data2 = data1[4].split("[")
                ff_usage = data2[0].strip()
                ff_percent = data2[1].strip(" %]")
                #print (ff_usage)
                #print (ff_percent)
                data2 = data1[5].split("[")
                ram_usage = data2[0].strip()
                ram_percent = data2[1].strip(" %]")
                #print (ram_usage)
                #print (ram_percent)
                data2 = data1[6].split("[")
                dsp_usage = data2[0].strip()
                dsp_percent = data2[1].strip(" %]")
                #print (dsp_usage)
                #print (dsp_percent)
            elif re.search('Kernel Frequency', line) != None :
                #print (line)
                data1 = line.split()
                fmax_usage = float(data1[2].strip())
                #print (str(fmax_usage))

    return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, fmax_usage)

def impl_report_i_parsing(impl_name, impl_name2):
    lut_usage = ""
    ff_usage = ""
    ram_usage = ""
    dsp_usage = ""
    lut_percent = ""
    ff_percent = ""
    ram_percent = ""
    dsp_percent = ""
    fmax_usage = 0.0
    index = 0
    if os.path.isfile(impl_name) == False :
        return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, fmax_usage)
    #print ("i:" +  impl_name)
    with open(impl_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('ALUTs', line) != None :
                #print (line)
                data1 = line.split(":")
                lut_usage = data1[1].strip().replace(",", "")
                #print (lut_usage)
            elif re.search('Registers', line) != None :
                #print (line)
                data1 = line.split(":")
                ff_usage = data1[1].strip().replace(",", "")
                #print (ff_usage)
            elif re.search('RAM blocks', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split(":")
                data2 = data1[1].split()
                #print (data2)
                ram_usage = data2[0].replace(",", "")
                ram_percent = data2[4]
                #print (ram_usage)
            elif re.search('DSP blocks', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split(":")
                data2 = data1[1].split()
                #print (data2)
                dsp_usage = data2[0].replace(",", "")
                dsp_percent = data2[4]
                #print (dsp_usage)
            elif re.search('Total', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split("|")
                #print (data1)
                data2 = data1[3].split()
                #print (data2)
                ram_usage = data2[0].strip()
                ram_percent = data2[1].lstrip('(').rstrip(')')
                #print (ram_usage)
                data2 = data1[4].split()
                #print (data2)
                dsp_usage = data2[0].strip()
                dsp_percent = data2[1].lstrip('(').rstrip(')')
                #print (dsp_usage)
            elif re.search('Kernel fmax', line) != None or re.search('Kernel Frequency', line) != None :
                #print (line)
                data1 = line.split(":")
                data2 = data1[1].strip().split()
                fmax_usage = float(data2[0].strip())
                #print (fmax_usage)

    return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, fmax_usage)

def hls_report_parsing(hls_name, hls_name2):
    lut_usage = 0.0
    ff_usage = 0.0
    ram_usage = 0.0
    dsp_usage = 0.0
    latency_usage = 0
    index = 0
    #print ("x:" +  hls_name)
    if os.path.isfile(hls_name) == False :
        return (lut_usage, ff_usage, ram_usage, dsp_usage, latency_usage)
    with open(hls_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('\* BRAM', line) != None :
                index = 1
            elif re.search('\* DSP', line) != None :
                index = 2
            elif re.search('\* FF', line) != None :
                index = 3
            elif re.search('\* LUT', line) != None :
                index = 4
            elif re.search('.*_kernel ', line) != None :
                data1 = line.split("|")
                if len(data1) != 4 :
                    continue
                #print (line)
                #print (index)
                if index == 1 :
                    ram_usage = float(data1[2])
                    index = 0
                    #print ("ram_usage = " + str(ram_usage))
                elif index == 2 :
                    dsp_usage = float(data1[2].strip())
                    index = 0
                    #print ("dsp_usage = " + str(dsp_usage))
                elif index == 3 :
                    ff_usage = float(data1[2].strip())
                    index = 0
                    #print ("ff_usage = " + str(ff_usage))
                elif index == 4 :
                    lut_usage = float(data1[2].strip())
                    index = 0
                    #print ("lut_usage = " + str(lut_usage))
    if os.path.isfile(hls_name2) == False :
        return (lut_usage, ff_usage, ram_usage, dsp_usage, latency_usage)
    #print ("x:" +  hls_name2)
    with open(hls_name2) as file:
        lineList = file.readlines()
        for line in lineList:
            data1 = line.split("|")
            if re.search('Latency', line) != None :
                index = 1
            elif index == 1 and len(data1) == 7:
                if data1[2].strip().isdigit() :
                    #print (data1[2])
                    latency_usage = int(data1[2])
                    #print ("latency_usage = " + str(latency_usage))
    return (lut_usage, ff_usage, ram_usage, dsp_usage, latency_usage)

def hls_report_i_parsing(hls_name, hls_name2):
    lut_percent = ""
    ff_percent = ""
    ram_percent = ""
    dsp_percent = ""
    latency_usage = ""
    index = 0
    #print ("i:" +  hls_name)
    if os.path.isfile(hls_name) == False :
        return (lut_percent, ff_percent, ram_percent, dsp_percent, latency_usage)
    with open(hls_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('ALUTs', line) != None :
                data = line.strip(' ;').split(";")
                #print (data)
                lut_percent = data[1].strip(" %")
            elif re.search('Dedicated logic registers', line) != None :
                data = line.strip(' ;').split(";")
                #print (data)
                ff_percent = data[1].strip(" %")
            elif re.search('Memory blocks', line) != None :
                data = line.strip(' ;').split(";")
                #print (data)
                ram_percent = data[1].strip(" %")
            elif re.search('DSP blocks', line) != None :
                data = line.strip(' ;').split(";")
                #print (data)
                dsp_percent = data[1].strip(" %")
            elif re.search('Total', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split("|")
                #print (data1)
                data2 = data1[3].split()
                #print (data2)
                ram_usage = data2[0].strip()
                ram_percent = data2[1].lstrip('(').rstrip(')')
                #print (ram_usage)
                #print (ram_percent)
                data2 = data1[4].split()
                #print (data2)
                dsp_usage = data2[0].strip()
                dsp_percent = data2[1].lstrip('(').rstrip(')')
                #print (dsp_usage)
                #print (dsp_percent)
    return (lut_percent, ff_percent, ram_percent, dsp_percent, latency_usage)
    
def hls_report3_i_parsing(hls_name3):
    lut_usage = ""
    ff_usage = ""
    ram_usage = ""
    dsp_usage = ""
    lut_percent = ""
    ff_percent = ""
    ram_percent = ""
    dsp_percent = ""
    index = 0
    #print ("i:" +  hls_name3)
    if os.path.isfile(hls_name3) == False :
        #print (lut_usage, ff_usage, ram_usage, dsp_usage)
        return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent)
    with open(hls_name3) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('Total', line) != None :
                index = 1
            elif index == 1 and re.search('data_percent', line) != None :
                index = 3
            elif index == 1 and re.search('data', line) != None :
                index = 2
            elif index == 2 :
                data1 = line.strip().lstrip('[').rstrip(']').split(",")
                index = 1
                #print (data1)
            elif index == 3 :
                data2 = line.strip().lstrip('[').rstrip(']').split(",")
                index = 0
                #print (data2)
        if len(data1) >= 4:
            lut_usage = data1[0].strip()
            ff_usage = data1[1].strip()
            ram_usage = data1[2].strip()
            dsp_usage = data1[3].strip()
        if len(data2) >= 4:
            lut_percent = data2[0].strip()
            ff_percent = data2[1].strip()
            ram_percent = data2[2].strip()
            dsp_percent = data2[3].strip()
        
    #print (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent)
    return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent)

def hls_report_x_parsing(hls_name, hls_name2):
    lut_usage = ""
    ff_usage = ""
    ram_usage = ""
    dsp_usage = ""
    lut_percent = ""
    ff_percent = ""
    ram_percent = ""
    dsp_percent = ""
    latency_usage = "" 
    index = 0

    #print ("x:" +  hls_name)
    if os.path.isfile(hls_name) == False :
        #print (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, latency_usage)
        return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, latency_usage)
    with open(hls_name) as file:
        lineList = file.readlines()
        data1 = []
        data3 = []
        for line in lineList:
            if re.search('Total', line) != None and re.search('\|', line) != None :
                data1 = line.strip(' |').split("|")
                #print (data1)
            #if re.search('Available', line) != None :
            #    data2 = line.strip(' |').split("|")
            if re.search('Utilization \(%\)', line) != None :
                data3 = line.strip(' |').split("|")
                #print (data3)
            if re.search('Latency', line) != None :
                index = 1
            elif index == 1 :
                data = line.split("|")
                if len(data) == 7 and (data[2].strip().isdigit() or data[2].strip() == "?") :
                    latency_usage = data[2].strip()
                    index = 0
                    #print ("latency_usage = " + str(latency_usage))
        #print (data1)
        #print (data2)
        #print (data3)
        if len(data1) >= 4:
            lut_usage = data1[4].strip()
            ff_usage = data1[3].strip()
            ram_usage = data1[1].strip()
            dsp_usage = data1[2].strip()
        if len(data3) >= 4:
            lut_percent = data3[4].strip()
            ff_percent = data3[3].strip()
            ram_percent = data3[1].strip()
            dsp_percent = data3[2].strip()
        
    #print (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, latency_usage)
    return (lut_usage, lut_percent, ff_usage, ff_percent, ram_usage, ram_percent, dsp_usage, dsp_percent, latency_usage)

def merlin_report_i_parsing(report_name, report_name2):
    lut_usage = 0.0
    ff_usage = 0.0
    ram_usage = 0.0
    dsp_usage = 0.0
    runtime = 0.0
    memory = 0.0
    ac_usage = 0
    index = 0
    #print ("merlin report:" +  report_name)
    if os.path.isfile(report_name) == False :
        #print (lut_usage, ff_usage, ram_usage, dsp_usage, ac_usage, runtime, memory)
        return (lut_usage, ff_usage, ram_usage, dsp_usage, ac_usage, runtime, memory)
    with open(report_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('hierarchy', line) != None :
                index = 1
            elif index == 1 and re.search('kernel', line) != None :
                data1 = line.split("|")
                #print (data1)
                if data1[3].strip() != "-" and data1[3].strip() != "" and data1[3].strip() != "???" :
                    ac_usage = int(data1[3].strip())
                lut_usage = data1[5].strip()
                if lut_usage == "" :
                    lut_usage = 0.0
                ff_usage = data1[6].strip()
                if ff_usage == "" :
                    ff_usage = 0.0
                ram_usage = data1[7].strip()
                if ram_usage == "" :
                    ram_usage = 0.0
                dsp_usage = data1[8].strip()
                if dsp_usage == "" :
                    dsp_usage = 0.0
                index = 0
                #print ("ac_usage = " + str(ac_usage))
    
    with open(report_name2) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('Total time', line) != None :
                words = line.split()
                #print (words)
                temp = float(words[2])
                if runtime > 0.0 and temp < runtime :
                    runtime = temp
                elif runtime == 0.0 :
                    runtime = temp
            elif re.search('Peak memory usage', line) != None :
                words = line.split()
                #print (words)
                temp = float(words[3])
                if memory > 0.0 and temp < memory :
                    memory = temp
                elif memory == 0.0 :
                    memory = temp

    #print (lut_usage, ff_usage, ram_usage, dsp_usage, ac_usage, runtime, memory)
    return (lut_usage, ff_usage, ram_usage, dsp_usage, ac_usage, runtime, memory)

def merlin_report_x_parsing(report_name):
    lut_usage = 0
    ff_usage = 0
    ram_usage = 0
    dsp_usage = 0
    runtime = 0.0
    memory = 0.0
    if os.path.isfile(report_name) == False :
        return (lut_usage, ff_usage, ram_usage, dsp_usage, runtime, memory)
    #print (report_name)
    with open(report_name) as file:
        lineList = file.readlines()
        for line in lineList:
            if re.search('Total time', line) != None :
                words = line.split()
                #print (words)
                runtime = float(words[2])
            elif re.search('Peak memory usage', line) != None :
                words = line.split()
                #print (words)
                memory = float(words[3])
            elif re.search('.*_kernel ', line) != None and re.search('%', line) != None :
                #print (line)
                data1 = line.split("|")
                data2 = data1[2].split("[")
                lut_usage = float(data2[1].lstrip().rstrip(" %]"))
                #print (lut_usage)
                data2 = data1[4].split("[")
                ff_usage = float(data2[1].lstrip().rstrip(" %]"))
                #print (ff_usage)
                data2 = data1[5].split("[")
                ram_usage = float(data2[1].lstrip().rstrip(" %]"))
                #print (ram_usage)
                data2 = data1[6].split("[")
                dsp_usage = float(data2[1].lstrip().rstrip(" %]"))
                #print (dsp_usage)

    return (lut_usage, ff_usage, ram_usage, dsp_usage, runtime, memory)

def resource_report_none(opt):
    lut_usage = 0.0
    ff_usage = 0.0
    ram_usage = 0.0
    dsp_usage = 0.0
    if opt == "fmax" :
        opt_usage = 0.0
    else :
        opt_usage = 0
    #print ("n:" +  resource_name)
    return (lut_usage, ff_usage, ram_usage, dsp_usage, opt_usage)
    
def merlincc_report_setup(dir):
    # get resource file path
    global impl_report, impl_report2, curr_fpga_design
    global hls_report, hls_report2, hls_report3
    global merlin_report, merlin_report2
    impl_report = ""
    hls_report = ""
    hls_report2 = ""
    hls_report3 = ""
    merlin_report = ""
    merlin_report2 = ""
    curr_fpga_design = ""
    for root,dirnames,filenames in os.walk(dir) :
        #print (root)
        #print (dirnames)
        if "Makefile" in filenames and root != dir :
            del dirnames[:]
            #print (dirnames)
            continue
        #print (dirnames)
        subdirs = root.split(os.sep)
        # impl/ under merlincc_report folder
        if "merlincc_report" in subdirs and "merlin_impl.log" in filenames and impl_report == "" :
            impl_report = os.path.join(root, "merlin_impl.log")
            #curr_fpga_design = "x"
            #print (impl_report)
        # impl/ under merlincc_report folder
        elif "acl_quartus_report.txt" in filenames and "merlin_bit.log" in filenames :
            impl_report = os.path.join(root, "acl_quartus_report.txt")
            curr_fpga_design = "i"
            #print (impl_report)
        elif "acl_quartus_report.txt" in filenames and "summary.log" in filenames :
            impl_report = os.path.join(root, "acl_quartus_report.txt")
            curr_fpga_design = "i"
            #print (impl_report)
        elif "merlin_bit.log" in filenames and impl_report == "" :
            impl_report = os.path.join(root, "merlin_bit.log")
        elif "merlin.log" in filenames and "Makefile" in filenames :
            impl_report = merlin_report2 = os.path.join(root, "merlin.log")
            if "merlin.rpt" in filenames and root == dir :
                hls_report = os.path.join(root, "merlin.log")
                hls_report2 = os.path.join(root, "merlin.rpt")
                merlin_report = os.path.join(root, "merlin.rpt")
                merlin_report2 = os.path.join(root, "merlin.log")
        # hls/ under merlincc_report folder
        elif "summary.json" in filenames :
            hls_report3 = os.path.join(root, "summary.json")
        elif "merlin.rpt" in filenames and "vendor.log" in filenames :
            hls_report = os.path.join(root, "vendor.log")
            hls_report2 = os.path.join(root, "merlin.rpt")
            merlin_report = os.path.join(root, "merlin.rpt")
            merlin_report2 = os.path.join(root, "vendor.log")
            #print (hls_report)
            #print (hls_report2)

        if "sdaccel_profile_summary.csv" in filenames :
            curr_fpga_design = "x"
        elif "profile.mon" in filenames :
            curr_fpga_design = "i"
        for file in filenames:
            if file.endswith(".xclbin") :
                curr_fpga_design = "x"
            elif file.endswith("*.aocx") :
                curr_fpga_design = "i"

    #print ("~" + impl_report + "~")
    #print ("~" + hls_report + "~")
    #print ("~" + hls_report2 + "~")
    #print ("~" + hls_report3 + "~")
    #print ("~" + curr_fpga_design + "~")
    if impl_report == "" and hls_report == "" and hls_report2 == "" and merlin_report == "" and merlin_report2 == "" :
        return False
    else :
        return True

def qor_flow_single_test(dir):
    #print ("QoR on " + dir)
    qorlist = range(22)
    #run_sim_profile_setup(dir)
    run_bit_profile_setup(dir)
    if curr_fpga_design == "x" :
        qorlist[19], qorlist[20] = run_bit_profile_x_parsing(run_bit_profile)
    elif curr_fpga_design == "i" :
        qorlist[19], qorlist[20] = run_bit_profile_i_parsing(run_bit_profile)
    else :
        qorlist[19], qorlist[20] = run_bit_profile_none()
    resource_report_setup(dir)
    #print (resource_report, resource_report2)
    #print (curr_fpga_design)
    #print (hls_report, hls_report2)
    #print (merlin_report)
    if curr_fpga_design == "x" :
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5] = resource_report_none("lat")
        qorlist[14], qorlist[15], qorlist[16], qorlist[17], qorlist[18] = resource_report_x_parsing(resource_report)
        qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13] = hls_report_parsing(hls_report, hls_report2)
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[6], qorlist[7] = merlin_report_x_parsing(merlin_report)
    elif curr_fpga_design == "i" :
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5] = resource_report_none("lat")
        qorlist[14], qorlist[15], qorlist[16], qorlist[17], qorlist[18] = resource_report_i_parsing(resource_report, resource_report2)
        qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13] = hls_report_parsing(hls_report, hls_report2)
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[6], qorlist[7] = merlin_report_i_parsing(merlin_report)
    else :
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5] = resource_report_none("lat")
        qorlist[14], qorlist[15], qorlist[16], qorlist[17], qorlist[18] = resource_report_none("fmax")
        if hls_report != "" :
            qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13] = hls_report_parsing(hls_report, hls_report2)
        else :
            qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13] = resource_report_none("lat")
        qorlist[6], qorlist[7] = run_bit_profile_none()
    qorlist[0] = get_test_case_name(dir)
    return (qorlist)

def get_test_case_name(dir):
    dir = dir.rstrip(os.sep)
    words = dir.split(os.sep)
    name = words[-1]
    if re.search('intel', words[-1]) != None or re.search('xilinx', words[-1]) != None :
        for i in range(len(words)-1, 0, -1) :
            name = words[i-1] + os.sep + name
            if re.search('intel', words[i-1]) == None and re.search('xilinx', words[i-1]) == None :
                break
    return (name)

def sort_report_list(list):
    #print (list)
    list = sorted(list, key=lambda report: report[0])
    #print (list)
    return (list)

def print_cc_report(list, mode, listtype):
    #curr_date = datetime.today()

    # write .rpt file
    report_name = "qor_" + curr_date.strftime("%Y%02m%02d") + ".rpt"
    f = open(report_name, mode)
    f.write("Date: " + curr_date.ctime() + "\n\n")
    if show_hls == True : 
        f.write('{0:<35}|{1:<69}|{2:<78}'.format(listtype, "Merlin", "HLS"))
    if show_impl == True :
        f.write('|{0:<66}'.format("Impl"))
    if show_onboard == True :
        f.write('|{0:<38}'.format("Onboard"))
    f.write('\n')
    if show_hls == True : 
        f.write('{0:^35}|{1:>6} {2:>6} {3:>6} {4:>6} {5:>10} {6:>11} {7:>11} {8:>6}|{9:>7} {10:>8} {11:>7} {12:>8} {13:>7} {14:>8} {15:>7} {16:>8} {17:>10}'.format("test_case_name", "LUT%", "FF%", "RAM%", "DSP%", "LAT", "runtime", "memory", "Status", "LUT", "LUT%", "FF", "FF%", "RAM", "RAM%", "DSP", "DSP%", "LAT"))
    if show_impl == True :
        f.write('|{0:>7} {1:>5} {2:>7} {3:>5} {4:>7} {5:>5} {6:>7} {7:>5} {8:>10}'.format("LUT", "LUT%", "FF", "FF%", "RAM", "RAM%", "DSP", "DSP%", "fmax (MHz)"))
    if show_onboard == True :
        f.write('|{0:>15} {1:>15} {2:^6}'.format("T-kernel(ms)", "T-pcie(ms)", "RunResult"))
    f.write('\n')
    for items in list :
        #print (items)
        if show_hls == True : 
            f.write('{0:35}|{1:>6} {2:>6} {3:>6} {4:>6} {5:10d} {6:11.2f} {7:11.2f} {8:6}|{9:>7} {10:>8} {11:>7} {12:>8} {13:>7} {14:>8} {15:>7} {16:>8} {17:>10}'.format(items[0], items[1], items[2], items[3], items[4], items[5], items[6], items[7], items[8], items[9], items[10], items[11], items[12], items[13], items[14], items[15], items[16], items[17]))
        if show_impl == True :
            f.write('|{0:>7} {1:>5} {2:>7} {3:>5} {4:>7} {5:>5} {6:>7} {7:>5} {8:10.3f}'.format(items[18], items[19], items[20], items[21], items[22], items[23], items[24], items[25], items[26]))
        if show_onboard == True :
            f.write('|{0:15.9f} {1:15.9f} {2:6}'.format(items[27], items[28], items[29]))
        f.write('\n')
    f.close()

    # write .csv file
    report_name = "qor_" + curr_date.strftime("%Y%02m%02d") + ".csv"
    f = open(report_name, mode)
    f.write("Date: " + curr_date.ctime() + "\n\n")
    f.write(listtype + ",")
    if show_hls == True : 
        f.write("Merlin,,,,,,,,HLS,,,,,,,,,")
    if show_impl == True :
        f.write("Impl,,,,,,,,,")
    if show_onboard == True :
        f.write("Onboard,,,")
    f.write('\n')
    if show_hls == True : 
        f.write("test_case_name,LUT%,FF%,RAM%,DSP%,LAT,runtime,memory,Status,LUT,LUT%,FF,FF%,RAM,RAM%,DSP,DSP%,LAT,")
    if show_impl == True :
        f.write("LUT,LUT%,FF,FF%,RAM,RAM%,DSP,DSP%,fmax (MHz),")
    if show_onboard == True :
        f.write("T-kernel(ms),T-pcie(ms),RunResult,")
    f.write('\n')
    for items in list :
        if show_hls == True : 
            for i in range(18) :
                f.write(str(items[i]) + ",")
        if show_impl == True :
            for i in range(18, 27) :
                f.write(str(items[i]) + ",")
        if show_onboard == True :
            for i in range(27, 30) :
                f.write(str(items[i]) + ",")
        f.write("\n")
    f.close()
    return (report_name)

def print_flow_report(list, mode):
    #curr_date = datetime.today()
    report_name = "qor_" + curr_date.strftime("%Y%02m%02d") + ".rpt"
    f = open(report_name, mode)
    f.write("Date: " + curr_date.ctime() + "\n\n")
    f.write('{:<22}|{:<69}|{:<54}|{:<78}|{:<38}\n'.format("", "Merlin", "HLS", "Impl", "Onboard"))
    f.write('{:^22}|{:>6} {:>6} {:>6} {:>6} {:>10} {:>11} {:>11} {:^6}|{:>12} {:>12} {:>12} {:>12} {:>10}|{:>14} {:>14} {:>14} {:>14} {:>10}|{:>15} {:>15} {:^6}\n'.format("test_case_name", "LUT%", "FF%", "RAM%", "DSP%", "LAT", "T-kernel(s)", "T-pcie(s)", "Status", "LUT%", "FF%", "RAM%", "DSP%", "LAT", "LUT", "FF", "RAM", "DSP", "fmax (MHz)", "T-kernel(s)", "T-pcie(s)", "Status"))
    for items in list :
        #print (items)
        f.write('{:22}|{:6.2f} {:6.2f} {:6.2f} {:6.2f} {:10d} {:11.2f} {:11.2f} {:6}|{:>12} {:>12} {:>12} {:>12} {:10d}|{:>14} {:>14} {:>14} {:>14} {:10.3f} {:15.9f} {:15.9f} {:6}\n'.format(items[0], items[1], items[2], items[3], items[4], items[5], items[6], items[7], items[8], items[9], items[10], items[11], items[12], items[13], items[14], items[15], items[16], items[17], items[18], items[19], items[20], items[21]))
    f.close()

def qor_cc_single_test(dir):
    global curr_fpga_design
    #print ("QoR on " + dir)
    qorlist = range(30)
    qorlist[5] = 0
    #qorlist = []
    #for i in range(23):
    #    qorlist.append("")
    #run_sim_profile_setup(dir)
    curr_fpga_design = ""
    runaccexe_retval = runaccexe_profile_setup(dir)
    #print (runaccexe_retval)
    if show_onboard == True :
        if curr_fpga_design == "x" :
            qorlist[27], qorlist[28] = run_bit_profile_x_parsing(run_bit_profile)
        elif curr_fpga_design == "i" :
            qorlist[27], qorlist[28] = run_bit_profile_i_parsing(run_bit_profile)
        else :
            qorlist[27], qorlist[28] = run_bit_profile_none()
        if qorlist[27] == 0.0 and qorlist[28] == 0.0 :
            qorlist[29] = "Failed"
        else :
            qorlist[29] = "Pass"
        for root,dirnames,filenames in os.walk(dir) :
            if "fail.o" in filenames :
                qorlist[29] = "Failed"
                break

    merlincc_retval = merlincc_report_setup(dir)
    #print (merlincc_retval)
    #print (curr_fpga_design)
    if merlincc_retval == False :
        return []

    # parse data for HLS
    if curr_fpga_design == "x" :
        qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13], qorlist[14], qorlist[15], qorlist[16], qorlist[17] = hls_report_x_parsing(hls_report, hls_report2)
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5], qorlist[6], qorlist[7] = merlin_report_i_parsing(merlin_report, merlin_report2)
        #print (qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5], qorlist[6], qorlist[7])
        #print ("x:{0}".format(qorlist))
    elif curr_fpga_design == "i" :
        qorlist[10], qorlist[12], qorlist[14], qorlist[16], qorlist[17]  = hls_report_i_parsing(hls_report, hls_report2)
        qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13], qorlist[14], qorlist[15], qorlist[16] = hls_report3_i_parsing(hls_report3)
        qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5], qorlist[6], qorlist[7] = merlin_report_i_parsing(merlin_report, merlin_report2)
        #print ("i:{0}".format(qorlist))
    else :
        if os.path.isfile(hls_report) == False :
            qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5] = resource_report_none("lat")
            qorlist[6], qorlist[7] = run_bit_profile_none()
            #print ("~:{0}".format(qorlist))
        else :
            with open(hls_report, 'r') as f :
                for line in f :
                    if re.search('aoc', line) != None :
                        curr_fpga_design = "i"
                        break
                    elif re.search('xocc', line) != None :
                        curr_fpga_design = "x"
                        break
            if curr_fpga_design == "x" :
                qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13], qorlist[14], qorlist[15], qorlist[16], qorlist[17] = hls_report_x_parsing(hls_report, hls_report2)
                qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5], qorlist[6], qorlist[7] = merlin_report_i_parsing(merlin_report, merlin_report2)
                #print (" x:{0}".format(qorlist))
            elif curr_fpga_design == "i" :
                qorlist[10], qorlist[12], qorlist[14], qorlist[16], qorlist[17]  = hls_report_i_parsing(hls_report, hls_report2)
                qorlist[9], qorlist[10], qorlist[11], qorlist[12], qorlist[13], qorlist[14], qorlist[15], qorlist[16] = hls_report3_i_parsing(hls_report3)
                qorlist[1], qorlist[2], qorlist[3], qorlist[4], qorlist[5], qorlist[6], qorlist[7] = merlin_report_i_parsing(merlin_report, merlin_report2)
                #print (" i:{0}".format(qorlist))

    if qorlist[6] == 0.0 and qorlist[7] == 0.0 :
        qorlist[8] = "Failed"
    else :
        qorlist[8] = "Pass"
    #if qorlist[17].isdigit() == True :
    #    if int(qorlist[17]) != 0 :
    #        merlin2hls = float(qorlist[5]) / float(qorlist[17])
    #else :
    #    merlin2hls = 1.0
    #if qorlist[5] != 0 and qorlist[17].isdigit() == True :
    #    hls2merlin = float(qorlist[17]) / float(qorlist[5])
    #else :
    #    hls2merlin = 1.0
    #if merlin2hls > 2.5 or hls2merlin > 2.5 :
    #    qorlist[8] = "Failed"

    if show_impl == True :
        if curr_fpga_design == "x" :
            qorlist[18], qorlist[19], qorlist[20], qorlist[21], qorlist[22], qorlist[23], qorlist[24], qorlist[25], qorlist[26] = impl_report_x_parsing(impl_report)
        elif curr_fpga_design == "i" :
            qorlist[18], qorlist[19], qorlist[20], qorlist[21], qorlist[22], qorlist[23], qorlist[24], qorlist[25], qorlist[26] = impl_report_i_parsing(impl_report, impl_report2)

    qorlist[0] = get_test_case_name(dir)
    #print (qorlist)
    return (qorlist)

def compare_int(item1, item2):
    value1 = 0
    value2 = 0
    if item1 != "" :
        value1 = int(item1)
    if item2 != "" :
        value2 = int(item2)
    return (value1 - value2)

def compare_float(item1, item2):
    value1 = 0.0
    value2 = 0.0
    if item1 != "" :
        value1 = float(item1)
    if item2 != "" :
        value2 = float(item2)
    return (value1 - value2)

def compare_in_csv_format(words1, words2):
    thelist = range(8)
    #print (words1)
    #print (words2)
    thelist[0] = words1[0]
    thelist[1] = compare_int(words1[5], words2[5])
    thelist[2] = compare_int(words1[18], words2[18])
    thelist[3] = compare_int(words1[20], words2[20])
    thelist[4] = compare_int(words1[22], words2[22])
    thelist[5] = compare_int(words1[24], words2[24])
    thelist[6] = compare_float(words1[26], words2[26])
    thelist[7] = compare_float(words1[27], words2[27])
    #print (thelist)
    return (thelist)

def print_compare_report(lists, title):
    #curr_date = datetime.today()
    report_name = "qor_cmp_" + curr_date.strftime("%Y%02m%02d") + ".csv"
    f = open(report_name, "a")
    f.write("Date: " + curr_date.ctime() + "\n")
    f.write(title + "\n")
    f.write("test_case_name,Est. LAT,LUT,FF,RAM,DSP,fmax (MHz),T-kernel(s)\n")
    for items in lists :
        for i in range(len(items)) :
            f.write(str(items[i]) + ",")
        f.write("\n")
    f.close()
    return (report_name)

def compare_cc_report(compare_file, to_be_compared_file):
    cmplist = []

    with open(to_be_compared_file) as file:
        lineList = file.readlines()
 
    with open(compare_file, 'r') as f :
        for currline in f :
            if re.search(',', currline) != None :
                currwords = currline.split(',');
                if currwords[0] != "Intel" and currwords[0] != "Xilinx" and currwords[0] != "test_case_name" :
                    #print (currline)
                    for line in lineList:
                        if re.search(',', line) != None :
                            words = line.split(',');
                            if words[0] == currwords[0] :
                                #print (line)
                                one_list = compare_in_csv_format(currwords, words)                         
                                if one_list != None and len(one_list) > 0 :
                                    #print (one_list)
                                    cmplist.append(one_list)
    return (cmplist)

def print_compare_excel(cmpfile, currentqor, lastweekqor, lastquarterqor):
    wb = Workbook()
    sheetcnt = 0

    ws = wb.create_sheet(index=sheetcnt)
    sheetcnt += 1
    ws.title = os.path.splitext(cmpfile)[0]
    csvreader = csv.reader(open(cmpfile, 'rb'), delimiter=',',quotechar='"')
    rowcnt = 1
    for row in csvreader :
        for col in range(len(row)) :
            _ = ws.cell(column=col+1, row=rowcnt, value=row[col])
        rowcnt += 1

    ws = wb.create_sheet(index=sheetcnt)
    sheetcnt += 1
    ws.title = os.path.splitext(currentqor)[0]
    csvreader = csv.reader(open(currentqor, 'rb'), delimiter=',',quotechar='"')
    rowcnt = 1
    for row in csvreader :
        for col in range(len(row)) :
            _ = ws.cell(column=col+1, row=rowcnt, value=row[col])
        rowcnt += 1

    if lastweekqor != "" :
        ws = wb.create_sheet(index=sheetcnt)
        sheetcnt += 1
        ws.title = os.path.splitext(lastweekqor)[0]
        csvreader = csv.reader(open(lastweekqor, 'rb'), delimiter=',',quotechar='"')
        rowcnt = 1
        for row in csvreader :
            for col in range(len(row)) :
                _ = ws.cell(column=col+1, row=rowcnt, value=row[col])
            rowcnt += 1

    if lastquarterqor != "" :
        ws = wb.create_sheet(index=sheetcnt)
        sheetcnt += 1
        ws.title = os.path.splitext(lastquarterqor)[0]
        csvreader = csv.reader(open(lastquarterqor, 'rb'), delimiter=',',quotechar='"')
        rowcnt = 1
        for row in csvreader :
            for col in range(len(row)) :
                _ = ws.cell(column=col+1, row=rowcnt, value=row[col])
            rowcnt += 1

    report_name = "qor_cmp_" + curr_date.strftime("%Y%02m%02d") + ".xlsx"
    wb.save(report_name)

def main():
    global show_hls, show_impl, show_onboard
    argvsize = len(sys.argv)
    if argvsize >= 2 :
        testlist = []
        ccilist = []
        ccxlist = []
        flowilist = []
        flowxlist = []
        lastweekqor = ""
        lastquarterqor = ""
        currentqor = ""
        parentdirname = ""
        for i in range(argvsize) :
            if sys.argv[i] == "hls_only" :
                show_hls = True
                show_impl = False
                show_onboard = False
            elif re.search('lastweek', sys.argv[i]) != None :
                words = sys.argv[i].split("=")
                lastweekqor = words[1]
                #print (lastweekqor)
            elif re.search('lastquarter', sys.argv[i]) != None :
                words = sys.argv[i].split("=")
                lastquarterqor = words[1]
                #print (lastquarterqor)
            elif re.search('current', sys.argv[i]) != None :
                words = sys.argv[i].split("=")
                currentqor = words[1]
                #print (currentqor)
        if currentqor == "" :
            searchpath = sys.argv[1].rstrip(os.sep)
            parentdirname = os.path.dirname(searchpath)
            if parentdirname == "" :
                parentdirname = "."
            basename = os.path.basename(searchpath)
        #print ("parentdirname=" + parentdirname + ",basename=" + basename)
        if len(parentdirname) > 0 :
            for file1 in os.listdir(parentdirname) :
                if basename != "." and file1 != basename :
                    continue
                for root,dirnames,filenames in os.walk(os.path.join(parentdirname,file1)) :
                    #print (root)
                    #print (dirnames)
                    #print (filenames)
                    if "merlincc_report" in os.path.basename(root).split() :
                        #print (root)
                        #print (dirnames)
                        #print (filenames)
                        testcasepath = os.path.dirname(root) + "/"
                        #print ("1:" + testcasepath)
                        one_list = qor_cc_single_test(testcasepath)
                        if len(one_list) > 0 :
                            if curr_fpga_design == "x" :
                                ccxlist.append(one_list)
                            elif curr_fpga_design == "i" :
                                ccilist.append(one_list)
                        if ".merlin_prj" in dirnames :
                            dirnames.remove('.merlin_prj')
                        if "merlincc_report" in dirnames :
                            dirnames.remove('merlincc_report')
                        
                    elif "report" in os.path.basename(root).split() :
                        for filename in fnmatch.filter(filenames, "merlin.log") :
                            #print (root.split("/"))
                            if ".merlin_prj" in root.split("/") :
                                continue
                            testcasepath = os.path.dirname(os.path.dirname(root)) + "/"
                            #print ("2:" + testcasepath)
                            one_list = qor_flow_single_test(testcasepath)
                            if len(one_list) > 0 :
                                if curr_fpga_design == "x" :
                                    flowxlist.append(one_list)
                                elif curr_fpga_design == "i" :
                                    flowilist.append(one_list)
                        if ".merlin_prj" in dirnames :
                            dirnames.remove('.merlin_prj')
                        if "merlincc_report" in dirnames :
                            dirnames.remove('merlincc_report')
                    elif "Makefile" in filenames :
                        #print ("it contains Makefile," + file1 + "\n")
                        #print (root)
                        #print (dirnames)
                        #print (filenames)
                        testcasepath = root + "/"
                        #print ("3:" + testcasepath)
                        one_list = qor_cc_single_test(testcasepath)
                        if len(one_list) > 0 :
                            if curr_fpga_design == "x" :
                                ccxlist.append(one_list)
                            elif curr_fpga_design == "i" :
                                ccilist.append(one_list)
                        if ".merlin_prj" in dirnames :
                            dirnames.remove('.merlin_prj')
                        if "merlincc_report" in dirnames :
                            dirnames.remove('merlincc_report')
                        
                    #else :
                    #    print ("\n")
                    #    print (root)
                    #    print (dirnames)
                    #    print (filenames)

                #if os.path.isdir(file1) and file1 != "report" :
                    #dirs2 = os.listdir(file1 + "/run/work/")
                    #for file2 in dirs2 :
                    #    if os.path.isdir(file1 + "/run/work/" + file2) :
                    #        print (file1 + "/run/work/" + file2)
                    #        testlist.append(qor_flow_single_test(file1 + "/run/work/" + file2 + "/"))
        if len(ccilist) > 0 :
            ccilist = sort_report_list(ccilist)
            currentqor = print_cc_report(ccilist, 'w', "Intel")
        if len(ccxlist) > 0 :
            ccxlist = sort_report_list(ccxlist)
            currentqor = print_cc_report(ccxlist, 'a', "Xilinx")
        if len(flowilist) > 0 :
            print_flow_report(flowilist, 'w')
        if len(flowxlist) > 0 :
            print_flow_report(flowxlist, 'a')
        if len(lastweekqor) > 0 :
            cmplist = compare_cc_report(currentqor, lastweekqor)
            cmpfile = ""
            if len(cmplist) > 0 :
                #print (cmplist)
                cmpfile = print_compare_report(cmplist, "Compare with last week")
        if len(lastquarterqor) > 0 :
            cmplist = compare_cc_report(currentqor, lastquarterqor)
            cmpfile = ""
            if len(cmplist) > 0 :
                #print (cmplist)
                cmpfile = print_compare_report(cmplist, "Compare with last release")
        if len(lastweekqor) > 0 or len(lastquarterqor) > 0 :
            print_compare_excel(cmpfile, currentqor, lastweekqor, lastquarterqor)
    else :
        print ("Usage: " + sys.argv[0] + " <test_case_directory>")

if __name__ == "__main__" :
    main()

