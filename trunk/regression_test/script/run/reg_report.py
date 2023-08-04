#(C) Copyright 2016-2021 Xilinx, Inc.
#All Rights Reserved.
#
#Licensed to the Apache Software Foundation (ASF) under one
#or more contributor license agreements.  See the NOTICE file
#distributed with this work for additional information
#regarding copyright ownership.  The ASF licenses this file
#to you under the Apache License, Version 2.0 (the
#"License"); you may not use this file except in compliance
#with the License.  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#Unless required by applicable law or agreed to in writing,
#software distributed under the License is distributed on an
#"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
#KIND, either express or implied.  See the License for the
#specific language governing permissions and limitations
#under the License. (edited)
#!/usr/bin/python

import sys
import os
import fnmatch
import re
import datetime

import os.path
import sys
import commands


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter

now_time = datetime.datetime.now().strftime('%Y-%m-%d')
print now_time
suite_total = 0
suite_name = []
total = []
failed = []
failed_list = []
comment = []
log_name = ""
dict_suite_name = {'unit_test': 'unit_test', 'commandline': 'commandline', 'bugzilla': 'bugzilla', 'cov_a': 'feature_cov', 'f_a_s': 'fcs_altera', 'm_a_s': 'machsuite_altera', 'c_a_s': 'chstone_altera', 'f_x': 'fcs_xilinx', 'm_x': 'machsuite_xilinx' }
dict_total = {'unit_test': 301, 'commandline': 28, 'bugzilla': 287, 'cov_a': 24, 'f_a_s': 16, 'm_a_s': 19, 'c_a_s': 10, 'f_x': 14, 'm_x': 18 }

# initial settings in the worksheet
def SheetSettings(ws):
    ws.title = sys.argv[2]
    ws['B2'] = 'Suite Name'
    ws['C2'] = 'Total'
    ws['D2'] = 'Failed'
    ws['E2'] = 'Pass Rate'
    ws['F2'] = 'Failed List'
    ws['G2'] = 'Comment'

    ws.column_dimensions['B'].width = 20.125
    ws.column_dimensions['C'].width = 6.5
    ws.column_dimensions['D'].width = 6.625
    ws.column_dimensions['F'].width = 20.25
    ws.column_dimensions['G'].width = 36.0

    ws.row_dimensions[2].height = 25.5
    ws.row_dimensions[14].height = 15.75
    ws.row_dimensions[15].height = 14.25
    ws.row_dimensions[17].height = 36.0

    AlignCells(ws, 'B2', 'G2', 'center')

    SetBgColorCells(ws, 'B2', 'G2', 'FF5072E0')

# clear content to range of cells
def ClearCells(ws, rows, cols):
    for row in ws[rows + ":" + cols] :
        for cell in row :
            cell.value = None

# Styling a Range or even for whole column are not implemented.
# You have to do it by your own
def AlignCells(ws, rows, cols, alignment):
    for row in ws[rows + ":" + cols] :
        for cell in row :
            cell.alignment = Alignment(horizontal=alignment,
                                       vertical=alignment)

# set solid background color to range of cells
def SetBgColorCells(ws, rows, cols, color):
    for row in ws[rows + ":" + cols] :
        for cell in row :
            cell.fill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type='solid')

# set borders to range of cells
def SetBorderCells(ws, rows, cols):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws[rows + ":" + cols] :
        for cell in row :
            cell.border = thin_border

# parse the summary log file
def ParseSummaryLog():
    global  suite_total, suite_name, total, failed, failed_list, comment
    global log_name
    for file in os.listdir('.') :
        if fnmatch.fnmatch(file, 'report_summary_*.log') :
            log_name = file
    with open(log_name) as file :
        lineList = file.readlines()
        for line in lineList:
            if re.search('Reporting .*?:', line) != None :
                suite_total += 1
                list = re.split(" |:", line)
                suite_name.append(list[1])
                total.append("")
                total[suite_total-1] = 0
                failed.append("")
                failed[suite_total-1] = 0
                failed_list.append("")
                comment.append("")
            elif re.search('\s.*?\sFail:Fail', line) != None :
                line = line.replace("Fail:Fail", "Fail")
                total[suite_total-1] += 1
                failed[suite_total-1] += 1
                failed_list[suite_total-1] += line.lstrip()
            elif re.search('\s.*?\sFail', line) != None :
                total[suite_total-1] += 1
                failed[suite_total-1] += 1
                failed_list[suite_total-1] += line.lstrip()

    #return suite_total, suite_name, total, failed, failed_list, comment

# set row height according to number of lines in multi-lines string
def SetRowHeight(ws, row, string):
    length = len(string.split('\n'))
    if length == 0 :
        length = 1
    ws.row_dimensions[row].height = (14.25 * length) + 2.0

def Combine_f_x():
    global  suite_total, suite_name, total, failed, failed_list, comment
    for i in range(suite_total) :
        if suite_name[i] == "f_x" :
            index_f_x = i
            break
    delete_list = []
    for i in range(len(suite_name)) :
        if re.search('f_x.*?', suite_name[i]) != None and i != index_f_x :
            list = re.split("\n", failed_list[i])
            for case in list :
                if re.search(case, failed_list[index_f_x]) == None :
                    total[index_f_x] += 1
                    failed[index_f_x] += 1
                    failed_list[index_f_x] += case + "\n"
            delete_list.append(i)
    for i in range(len(delete_list)) :
        del suite_name[delete_list[i]]
        del total[delete_list[i]]
        del failed[delete_list[i]]
        del failed_list[delete_list[i]]
        suite_total -= 1

def Combine_m_x():
    global  suite_total, suite_name, total, failed, failed_list, comment
    for i in range(suite_total) :
        if suite_name[i] == "m_x" :
            index_m_x = i
            break
    delete_list = []
    for i in range(len(suite_name)) :
        if re.search('m_x.*?', suite_name[i]) != None and i != index_m_x :
            list = re.split("\n", failed_list[i])
            for case in list :
                if re.search(case, failed_list[index_m_x]) == None :
                    total[index_m_x] += 1
                    failed[index_m_x] += 1
                    failed_list[index_m_x] += case + "\n"
            delete_list.append(i)
    for i in range(len(delete_list)) :
        del suite_name[delete_list[i]]
        del total[delete_list[i]]
        del failed[delete_list[i]]
        del failed_list[delete_list[i]]
        suite_total -= 1

def Skip_merlin_flow():
    global  suite_total, suite_name, total, failed, failed_list, comment
    for i in range(suite_total) :
        if suite_name[i] == "merlin_flow" :
            del suite_name[i]
            del total[i]
            del failed[i]
            del failed_list[i]
            suite_total -= 1
            break

def CreateNewExcelWorksheet():
    # open the excel file
    if os.path.isfile(sys.argv[1]) == True :
        wb = load_workbook(filename = sys.argv[1])
    else :
        wb = Workbook()
    wb_sheets = len(wb._sheets)
    # print (wb_sheets)

    # open the sheet
    ws = wb.active
    if wb_sheets == 1 :
        # with new created worksheet, set settings to it
        SheetSettings(ws)
    else :
        if ws.title != sys.argv[2] :
            # create and set settings to a new worksheet at index=0 position
            refws = ws
            ws = wb.create_sheet(index=0)
            SheetSettings(ws)
            # copy worksheet from specific existed worksheet
            # ws = wb.copy_worksheet(ws)
            # ws.title = sys.argv[2]
            # wb._sheets.sort(key=lambda ws: ws.title, reverse=True)
            # print (wb.index(ws))
        else :
            ClearCells(ws, 'C3', 'D11')
            ClearCells(ws, 'F3', 'G11')

        # print ( "'B'=" + str(ws.column_dimensions['B'].width) )
        # print ( "'C'=" + str(ws.column_dimensions['C'].width) )
        # print ( "'D'=" + str(ws.column_dimensions['D'].width) )
        # print ( "'E'=" + str(ws.column_dimensions['E'].width) )
        # print ( "'F'=" + str(ws.column_dimensions['F'].width) )
        # print ( "'G'=" + str(ws.column_dimensions['G'].width) )
        # print ( "'2'=" + str(ws.row_dimensions[2].height) )
        # print ( "'14'=" + str(ws.row_dimensions[14].height) )
        # print ( "'15'=" + str(ws.row_dimensions[15].height) )
        # print ( "'16'=" + str(ws.row_dimensions[16].height) )
        # print ( "'17'=" + str(ws.row_dimensions[17].height) )
        # print ( ws['B2'].fill )
        # print ( ws['B2'].fill.end_color.index )
        # print ( ws['B14'].font)
        # index_color = refws['B2'].fill.start_color.index
        # print (index_color)
        # index_color = refws['B3'].fill.start_color.index
        # print (index_color)

    # parse data from summary log file
    ParseSummaryLog()
    Combine_f_x()
    Combine_m_x()
    Skip_merlin_flow()
    # fill in parse result into table
    total_total = 0
    failed_total = 0
    for i in range(suite_total) :
        # fill in 'suite name'
        _ = ws.cell(column=2, row=i+3, value=dict_suite_name[suite_name[i]])
        # fill in 'total'
        if dict_total[suite_name[i]] < failed[i] :
            total[i] = failed[i]
        else :
            total[i] = dict_total[suite_name[i]]
        total_total += total[i]
        _ = ws.cell(column=3, row=i+3, value=total[i])
        # fill in 'failed'
        failed_total += failed[i]
        _ = ws.cell(column=4, row=i+3, value=failed[i])
        # fill in 'pass rate'
        ws['E'+str(i+3)].number_format = '0.00%'
        pass_rate = float(total[i] - failed[i]) / total[i]
        _ = ws.cell(column=5, row=i+3, value=pass_rate)
        # fill in 'failed list'
        failed_list[i] = failed_list[i].rstrip()
        SetRowHeight(ws, i+3, failed_list[i])
        ws['F'+str(i+3)].alignment = Alignment(wrap_text=True)
        _ = ws.cell(column=6, row=i+3, value=failed_list[i])

    # final total, failed, pass rate
    _ = ws.cell(column=3, row=suite_total+3, value=total_total)
    _ = ws.cell(column=4, row=suite_total+3, value=failed_total)
    ws['E'+str(suite_total+3)].number_format = '0.00%'
    if total_total != 0 :
        pass_rate = float(total_total - failed_total) / total_total
    else :
        pass_rate = 0
    _ = ws.cell(column=5, row=suite_total+3, value=pass_rate)

    # Cells' alignment, table cell borders
    AlignCells(ws, 'B3', 'E'+str(suite_total+2), 'center')
    SetBorderCells(ws, 'B2', 'G'+str(suite_total+2))

    # display build related information
    ws.row_dimensions[suite_total+5].height = 15.75
    ws.row_dimensions[suite_total+6].height = 14.25
    ws.row_dimensions[suite_total+7].height = 36.0
    ws.merge_cells('B'+str(suite_total+5)+':D'+str(suite_total+8))
    ws['B'+str(suite_total+5)].alignment = Alignment(wrap_text=True)
    ws['B'+str(suite_total+5)].font = Font(name='Calibri', size=16, color='FF00B050')
    ws['B'+str(suite_total+5)] = "Merlin build date:"+ now_time + "\nXOCC: 2017.1\nAOCL: 17.0.2.297\nTest_Env_Revision: 9451"

    # ws = wb.get_sheet_by_name(sys.argv[2])
    # wb.remove_sheet(ws)

    wb.save(sys.argv[1])

def SaveAsExcelWorksheet():
    # your starting wb with 2 Sheets: Sheet1 and Sheet2
    wb = load_workbook(filename = sys.argv[1])
    sheets = wb.sheetnames # ['Sheet1', 'Sheet2']

    for s in sheets:
        if s != sys.argv[2] :
            sheet_name = wb.get_sheet_by_name(s)
            wb.remove_sheet(sheet_name)

    # your final wb with just Sheet1
    wb.save('temporary_excel_for_html.xlsx')

def SaveAsHtmlWorksheet():
    os.system("libreoffice --headless --convert-to html temporary_excel_for_html.xlsx")
    outputList = []
    bAppend = False
    with open("temporary_excel_for_html.html") as file :
        lineList = file.readlines()
        for line in lineList :
            if re.search('<table', line) != None :
                bAppend = True
            elif re.search('</table', line) != None :
                bAppend = False
                outputList.append(line)
            if bAppend == True :
                outputList.append(line)
    file = open("temporary_excel_for_html.html", "w")
    file.writelines(outputList)
    file.close()


def ComposeReportToMaillist():
    f = open("temporary_excel_for_html.txt", 'w')
    if os.path.isfile(sys.argv[3]) == True :
        f.write("To: ")
        with open (sys.argv[3]) as inf :
            lines = inf.readlines()
            f.writelines(lines)
    else :
        f.write("To: " + sys.argv[3])
    f.write("\nSubject: regression result\n")
    f.write("MIME-Version: 1.0\n")
    f.write("Content-Type: multipart/mixed; boundary=asdfghjkl\n")
    f.write("\n--asdfghjkl\n")
    f.write("Content-Type: text/html\n\n")
    f.write("<pre class=\"western\">\n")
    with open (log_name) as inf :
        lines = inf.readlines()
        f.writelines(lines)
    f.write("\n</pre>\n")
    with open ("temporary_excel_for_html.html") as inf :
        lines = inf.readlines()
        f.writelines(lines)
    f.write("\n\n--asdfghjkl\n")
    f.write("Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheeti;\n")
    f.write("Content-Disposition: attachment; filename=\"" + sys.argv[1] + "\";\n")
    f.write("Content-Transfer-Encoding: base64\n\n")
    os.system("base64 " + sys.argv[1] + ">  temporary_excel_for_html.base64")
    with open ("temporary_excel_for_html.base64") as inf :
        lines = inf.readlines()
        f.writelines(lines)
    f.write("\n\n--asdfghjkl--\n")
    f.close()

def MailReportToMaillist():
    os.system("sendmail -i -t < temporary_excel_for_html.txt")

def TeardownReportToMaillist():
    os.system("rm -f temporary_excel_for_html.*")

def main():
    if len(sys.argv) == 4 :
        CreateNewExcelWorksheet()
        SaveAsExcelWorksheet()
        SaveAsHtmlWorksheet()
        ComposeReportToMaillist()
        MailReportToMaillist()
        TeardownReportToMaillist()
    else :
        print ("Usage: <app_name> <excel_file_name> <sheet_name> <maillist>")

if __name__ == "__main__" :
    main()
